#!/usr/bin/env python3
"""
Resident 360 / Visitor Dashboard — daily VMS cleaner.

Scans each project's raw VMS file and rebuilds the dashboard's competitor
dataset (repo file: visitor.xlsx, sheet FINAL).

Rules (per CK, Jul 2026):
  - Keep ONLY these Check In Purposes:
      Laundry, Inspection, Cleaning/Cleaners, Maintenance/Handyman,
      Fit-Out, AMC Contractors  (matched space/case-insensitively)
  - Keep ONLY rows where Check In Type == "Unit Visit".
  - Building/ Unit and Project Name are taken AS-IS from the source file
    (each source Excel maintains its own clean formula) — no derivation here.
  - THE8 file is per-person: collapse duplicate people so one
    (date + purpose + unit + company) counts once. Other files are per-visit.
  - Company Name is upper-cased (cosmetic; dashboard matches case-insensitively
    and strips our own companies Candoo / S&C / Strive on its side).
  - Dates: accept real Excel dates AND text dates in DDMMYYYY[ HH:MM:SS] form
    (e.g. NORTH). Rows with unparseable or implausible dates (year <2024 or
    >2027, e.g. a 4670 typo) are dropped.

Output schema (sheet FINAL), matching the existing repo file:
  Check In Date | Check In Type | Check In Purpose | Company Name |
  Scope of work | Building/ Unit | Project Name

Run:  python3 clean_vms.py
"""
import openpyxl, datetime, sys, os, re, glob

KEEP = {"LAUNDRY", "INSPECTION", "CLEANING/CLEANERS",
        "MAINTENANCE/HANDYMAN", "FIT-OUT", "AMCCONTRACTORS"}
norm = lambda s: str(s or '').upper().replace(' ', '')


def resolve(rel):
    """Find a file/folder under any session mount by its OneDrive-relative path.
    The mount base (/sessions/<name>/mnt) changes every session, so we glob."""
    hits = glob.glob('/sessions/*/mnt/' + rel)
    if hits:
        return hits[0]
    # Fallback: exact path if someone runs it with a fixed mount
    return '/sessions/current/mnt/' + rel


# project, OneDrive-relative path, sheets(None=all), dedupe_per_person
SOURCES = [
 ('BALQIS RESIDENCE',
  "Balqis Security's files - DAILY CONTRACTORS RECORDS/Visitors Details Balqis Residence.xlsx",
  None, False),
 ('THE8',
  "Abdul  Muqeet's files - VMS-DATA FILES/VMS-TH8.xlsx",
  ["VMS-TH8-'26"], True),
 ('NORTH RESIDENCE',
  "VMS DATA SOUTH & NORTH/VISITOR DATA NORTH RESIDENCE - 2026.xlsx",
  ['NORTH RESIDENCE'], False),
 ('SOUTH RESIDENCE',
  "VMS DATA SOUTH & NORTH/VISITOR DATA SOUTH RESIDENCE - 2026.xlsx",
  ['SOUTH RESIDENCE'], False),
]

OUT = resolve("CRM Related/Visitor-Competitor-Dashboard/visitor.xlsx")
HEADER = ['Check In Date', 'Check In Type', 'Check In Purpose', 'Company Name',
          'Scope of work', 'Building/ Unit', 'Project Name']


def col(ci, *names):
    low = {k.lower(): v for k, v in ci.items()}
    for n in names:
        if n.lower() in low:
            return low[n.lower()]
    return None


def parse_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        dt = v if isinstance(v, datetime.datetime) else datetime.datetime(v.year, v.month, v.day)
    elif isinstance(v, str) and v.strip():
        m = re.match(r'^(\d{2})(\d{2})(\d{4})(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?)?', v.strip())  # DDMMYYYY
        if not m:
            return None
        dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        hh = int(m.group(4) or 0); mi = int(m.group(5) or 0); ss = int(m.group(6) or 0)
        try:
            dt = datetime.datetime(yy, mm, dd, hh, mi, ss)
        except ValueError:
            return None
    else:
        return None
    if dt.year < 2024 or dt.year > 2027:
        return None
    return dt


def clean_source(proj, path, sheets, dedupe):
    if not os.path.exists(path):
        print(f'  !! MISSING: {proj} -> {path}', file=sys.stderr)
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sns = sheets or wb.sheetnames
    out, seen, dropped = [], set(), 0
    for sn in sns:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        hdr = None
        for row in it:
            vals = [str(c) if c is not None else '' for c in row]
            if 'Check In Date' in vals:
                hdr = vals
                break
        if not hdr:
            continue
        ci = {h: i for i, h in enumerate(hdr)}
        iD = col(ci, 'Check In Date'); iT = col(ci, 'Check In Type'); iP = col(ci, 'Check In Purpose')
        iC = col(ci, 'Company Name', 'COMPANY NAME'); iS = col(ci, 'Scope of work')
        iBU = col(ci, 'Building/ Unit', 'Building/Unit'); iPN = col(ci, 'Project Name', 'Project name')
        iU = col(ci, 'Unit')
        for row in it:
            if not row or (iD is not None and row[iD] in (None, '')):
                continue
            if norm(row[iP] if iP is not None else '') not in KEEP:
                continue
            typ = str(row[iT]).strip().lower() if iT is not None and row[iT] is not None else ''
            if typ != 'unit visit':
                continue
            date = parse_date(row[iD])
            if date is None:
                dropped += 1
                continue
            pur = str(row[iP]).strip()
            comp = str(row[iC]).strip().upper() if iC is not None and row[iC] is not None else ''
            scope = str(row[iS]).strip() if iS is not None and row[iS] is not None else ''
            bu = str(row[iBU]).strip() if iBU is not None and row[iBU] is not None else ''
            pn = str(row[iPN]).strip() if iPN is not None and row[iPN] is not None else proj
            unit = str(row[iU]).strip() if iU is not None and row[iU] is not None else ''
            if dedupe:
                key = (date.strftime('%Y-%m-%d'), norm(pur), unit.upper(), comp)
                if key in seen:
                    continue
                seen.add(key)
            out.append([date, 'Unit Visit', pur, comp, scope, bu, pn])
    print(f'  {proj}: {len(out)} rows (dropped {dropped} bad-date)')
    return out


def main():
    rows = []
    print('Cleaning sources:')
    for proj, path, sheets, dedupe in SOURCES:
        rows += clean_source(proj, path, sheets, dedupe)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FINAL'
    ws.append(HEADER)
    for r in rows:
        ws.append(r)
    wb.save(OUT)
    print(f'\nWROTE {len(rows)} rows -> {OUT}')
    return len(rows)


if __name__ == '__main__':
    main()
